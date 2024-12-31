# temporal network uses improved algorithms

import igraph as ig
import xlrd
import xlwt
import numpy as np
from collections import defaultdict
from random import choice
import pandas as pd
from collections import Counter
import random
from bisect import bisect_right
# from sets import Set
import sys  as _sys
import collections as _co
import openpyxl
import networkx as nx


class TemporalNetwork(ig.Graph):

    def __init__(self):
        ig.Graph.__init__(self)
        self.network = ig.Graph(1)
        self.tedges = []
        self.nodes = []
        self.time = defaultdict(lambda: list())
        self.targets = defaultdict(lambda: dict())
        self.sources = defaultdict(lambda: dict())
        self.activities = defaultdict(lambda: list())
        self.activities_sets = defaultdict(lambda: set())
        self.paths = _co.defaultdict(lambda: _co.defaultdict(lambda: np.array([0, 0])))
        self.ordered_times = []

        for e in self.tedges:
            self.time[e[2]].append(e)
            self.targets[e[2]].setdefault(e[1], []).append(e)
            self.sources[e[2]].setdefault(e[0], []).append(e)
            if e[0] not in self.nodes:
                self.nodes.append(e[0])
            if e[1] not in self.nodes:
                self.nodes.append(e[1])
        self.ordered_times = sorted(self.time.keys())

    def addEdge(self, source, target, start_time, end_time):
        e = (source, target, start_time, end_time)
        self.tedges.append(e)
        if source not in self.nodes:
            self.nodes.append(source)
        if target not in self.nodes:
            self.nodes.append(target)

        self.time[start_time].append(e)

        self.ordered_times = sorted(self.time.keys())

    def times_nodes(self, start_time, end_time):
        time_nodes = []
        for x in self.tedges:
            if start_time <= x[2] <= end_time:
                if x[1] not in time_nodes:
                    time_nodes.append(x[1])
                if x[0] not in time_nodes:
                    time_nodes.append(x[0])
        return time_nodes

    def time_edges(self, start_time, end_time):
        time_edges = 0
        for t in self.ordered_times:
            if start_time <= t <= end_time:
                time_edges += 1
        return time_edges

    def getObservationLength(self):
        return max(self.ordered_times) - min(self.ordered_times)

    def getInterEventTimes(self):
        timediffs = []
        for i in range(1, len(self.ordered_times)):
            timediffs += [self.ordered_times[i] - self.ordered_times[i - 1]]
        return np.array(timediffs)

    def Summary(self):
        summary = ''
        summary += 'Nodes number:\t' + str(len(self.nodes)) + '\n'
        summary += 'Edges number:\t' + str(len(self.tedges)) + '\n'
        if len(self.ordered_times) > 0:

            summary += 'Observation period:[' + str(min(self.ordered_times)) + ', ' + str(
                max(self.ordered_times)) + ']\n'
            summary += 'Time stamps:\t\t' + str(len(self.ordered_times)) + '\n'  #

        return summary

    def randomWalk(self, vlist, stop,cycle_time):
        l1 = [d[0] for d in vlist]
        t_list = []
        for ct in range(cycle_time):
            node = choice(l1)
            t_list.append(node)
            t = 1
            t1 = 0
            while (t <= stop):
                j = 0
                beixuan = [([0] * 3) for i in range(10000)]
                neighbors = []
                for i in range(len(vlist)):

                    if vlist[i][0] == node:
                        if vlist[i][2] > t1:
                            beixuan[j][0] = vlist[i][1]
                            beixuan[j][1] = vlist[i][2]
                            neighbors.append(vlist[i][1])
                            j = j+1
                            i = i+1
                        else:
                            i = i+1
                    else:
                        i = i+1
                print(len(neighbors))
                if len(neighbors) != 0:
                    node = choice(neighbors)
                    for i in range(len(beixuan)):
                        if beixuan[i][0] == node:
                            t1 = beixuan[i][1]
                    t_list.append(node)
                    t = t+1
                    beixuan.clear()
                    neighbors.clear()
                else:
                    print("no neighbors")
                    break
            ct = ct+1

        lista = dict(Counter(t_list))
        print(lista)

        key = list(lista.keys())
        value = list(lista.values())

        result_excel = pd.DataFrame()
        result_excel["company"] = key
        result_excel["times"] = value

        result_excel.to_excel('sx2-10000.xlsx')


    def expandSubPaths(self,vlist):
        for l in range(len(vlist)):
            path = (vlist[l][0],vlist[l][1])
            frequency = 1
            self.paths[1][path] += (0,frequency)
        for pathLength in range(0,100):
            for path in self.paths[pathLength]:
                frequency = self.paths[pathLength][path][1]
                for k in range(0, pathLength):
                    for s in range(0, pathLength - k + 1):
                        self.paths[k][path[s:s + k + 1]] += (frequency, 0)

    def getShortestPaths(self):
        shortest_paths = _co.defaultdict(lambda: _co.defaultdict(lambda: set()))
        shortest_path_lengths = _co.defaultdict(lambda: _co.defaultdict(lambda: np.inf))

        for l in self.paths:
            for p in self.paths[l]:
                s = p[0]
                d = p[-1]
                if l < shortest_path_lengths[s][d]:
                    shortest_path_lengths[s][d] = l
                    shortest_paths[s][d] = set()
                    shortest_paths[s][d].add(p)
                elif l == shortest_path_lengths[s][d]:
                    shortest_paths[s][d].add(p)
        return shortest_paths

    def BetweennessCentrality(self, normalized=False):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Sheet")
        node_centralities = _co.defaultdict(lambda: 0)
        shortest_paths = self.getShortestPaths()

        for s in shortest_paths:
            for d in shortest_paths[s]:
                for p in shortest_paths[s][d]:
                    for x in p:
                        if s != d != x:
                            node_centralities[x] += 1.0 / len(shortest_paths[s][d])
        m = 0
        for v in node_centralities:
            m += node_centralities[v]
        for v in node_centralities:
            node_centralities[v] /= m

        n = 2
        for v in self.nodes:
            node_centralities[v] += 0
            ws.cell(row=1, column=1).value = "nodes"
            ws.cell(row=1, column=2).value = "t-b"
            ws.cell(row=n, column=1).value = v
            ws.cell(row=n, column=2).value = node_centralities[v]
            n = n+1

        wb.save("t-b.xlsx")
        print("success")
