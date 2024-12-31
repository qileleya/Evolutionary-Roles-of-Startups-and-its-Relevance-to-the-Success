# static network uses improved algorithms

import igraph as ig
import xlrd
from random import choice
from collections import Counter
import pandas as pd
import xlwt
import openpyxl
import collections as _co
import numpy as np


class randNetwk(ig.Graph):  # random network inherits from Graph
    def __init__(self):
        ig.Graph.__init__(self)
        self.network = ig.Graph(1)
        self.paths = _co.defaultdict(lambda: _co.defaultdict(lambda: np.array([0, 0])))
        self.edges = []

    def ad_vertices(self, v_list):
        for i in range(len(v_list)):
            self.network.add_vertex(v_list[i])

    def ad_edges(self, e_list):
        for i in range(len(e_list)):
            self.network.add_edge(source=e_list[i][0], target=e_list[i][1])

    def randomWalk(self, vlist, stop,cycle_time):
        l1 = [d[0] for d in vlist]
        t_list = []
        for ct in range(cycle_time):
            node = choice(l1)
            t_list.append(node)
            t = 1
            t1 = 0
            while (t <= stop):
                j = 0
                beixuan = [([0] * 3) for i in range(10000)]
                neighbors = []
                for i in range(len(vlist)):

                    if vlist[i][0] == node:
                        if vlist[i][2] > t1:
                            beixuan[j][0] = vlist[i][1]
                            beixuan[j][1] = vlist[i][2]
                            neighbors.append(vlist[i][1])
                            j = j+1
                            i = i+1
                        else:
                            i = i+1
                    else:
                        i = i+1
                print(len(neighbors))
                if len(neighbors) != 0:
                    node = choice(neighbors)
                    for i in range(len(beixuan)):
                        if beixuan[i][0] == node:
                            t1 = beixuan[i][1]
                    t_list.append(node)
                    t = t+1
                    beixuan.clear()
                    neighbors.clear()
                else:
                    print("no neighbors")
                    break
            ct = ct+1

        lista = dict(Counter(t_list))
        print(lista)

        key = list(lista.keys())
        value = list(lista.values())

        result_excel = pd.DataFrame()
        result_excel["company"] = key
        result_excel["times"] = value

        result_excel.to_excel('j2-10000.xlsx')

    def expandSubPaths(self,vlist):
        for l in range(len(vlist)):
            path = (vlist[l][0],vlist[l][1])
            frequency = 1
            self.paths[1][path] += (0,frequency)
        for pathLength in range(0,100):
            for path in self.paths[pathLength]:
                frequency = self.paths[pathLength][path][1]
                for k in range(0, pathLength):
                    for s in range(0, pathLength - k + 1):
                        self.paths[k][path[s:s + k + 1]] += (frequency, 0)

    def getShortestPaths(self):
        shortest_paths = _co.defaultdict(lambda: _co.defaultdict(lambda: set()))
        shortest_path_lengths = _co.defaultdict(lambda: _co.defaultdict(lambda: np.inf))

        for l in self.paths:
            for p in self.paths[l]:
                s = p[0]
                d = p[-1]
                if l < shortest_path_lengths[s][d]:
                    shortest_path_lengths[s][d] = l
                    shortest_paths[s][d] = set()
                    shortest_paths[s][d].add(p)
                elif l == shortest_path_lengths[s][d]:
                    shortest_paths[s][d].add(p)
        return shortest_paths

    def BetweennessCentrality(self, normalized=False):
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("Sheet")
        node_centralities = _co.defaultdict(lambda: 0)
        shortest_paths = self.getShortestPaths()

        for s in shortest_paths:
            for d in shortest_paths[s]:
                for p in shortest_paths[s][d]:
                    for x in p:
                        if s != d != x:
                            node_centralities[x] += 1.0 / len(shortest_paths[s][d])
        m = 0
        for v in node_centralities:
            m += node_centralities[v]
        for v in node_centralities:
            node_centralities[v] /= m

        n = 2

        for v in range(len(self.network.vs)):
            node_centralities[self.network.vs[v]["name"]] += 0
            ws.cell(row=1, column=1).value = "node"
            ws.cell(row=1, column=2).value = "t-b"
            ws.cell(row=n, column=1).value = self.network.vs[v]["name"]
            ws.cell(row=n, column=2).value = node_centralities[self.network.vs[v]["name"]]
            n = n+1

        wb.save("static_improved.xlsx")
        print("success")
